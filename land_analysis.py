"""
地块核查数据分析工具 — 批量处理所有村组
从 "村数据" sheet 提取各地块的多维数据，按任务名称（村/社区）分组汇总，
匹配 Sheet1 的总亩数，生成完整的分析报表，并对比本地上报数据。

分析维度：
  1. 农作物种植用地属性
  2. 2025年夏收主要作物
  3. 2025年秋收主要作物

使用方法：
    python land_analysis.py <输入Excel>

示例：
    python land_analysis.py 地块核查导出数据到村20260515.xlsx

输出：在输入文件同目录生成 <输入文件名>_分析结果.xlsx
"""

import sys, os, re, time
from collections import Counter, defaultdict
import zipfile
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ── 样式常量 ──
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
VILLAGE_FILL = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
TASK_FILL = PatternFill("solid", fgColor="DAEEF3")
DIM_FILL = PatternFill("solid", fgColor="FFF2CC")
AREA_FILL = PatternFill("solid", fgColor="E2EFDA")
POS_FILL = PatternFill("solid", fgColor="C6EFCE")     # 正值（浅绿色）
NEG_FILL = PatternFill("solid", fgColor="FFC7CE")     # 负值（浅红色）

# ── 分析维度定义：(key, label, col_index_0based) ──
DIMENSIONS = [
    ("crop_attr", "农作物种植用地属性", 6),
    ("summer_crop", "2025年夏收主要作物", 7),
    ("autumn_crop", "2025年秋收主要作物", 12),
]

# ── 对照表：我方分类名 → 本地对比数据列名 ──
COMPARISON_MAP = {
    # 种植用地属性
    "旱地": "旱地",
    "其他水田": "其他水田",
    "水浇地": "水浇地",
    "冬水田":"冬水田",
    # 夏收作物
    "小麦": "小麦",
    "油菜": "油菜籽",
    "马铃薯": "马铃薯",
    "蔬菜": "蔬菜（含菜用瓜）",
    "其他经济作物": "中草药材",
    # 秋收作物
    "玉米": "玉米",
    "水稻": "稻谷",
    "大豆": "大豆",
    "甘薯": "甘薯",
    "高粱":"高粱",
}


def fix_excel_autofilter(excel_path: str) -> str:
    """修复因自动筛选 XML 损坏导致 openpyxl 无法读取的问题"""
    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix="_fixed.xlsx", delete=False)
    fixed_path = tmp.name
    tmp.close()
    with zipfile.ZipFile(excel_path, 'r') as zin:
        with zipfile.ZipFile(fixed_path, 'w') as zout:
            for item in zin.namelist():
                content = zin.read(item)
                if item.startswith('xl/worksheets/') and item.endswith('.xml'):
                    text = content.decode('utf-8', errors='replace')
                    text = re.sub(r'<autoFilter[^>]*>.*?</autoFilter>', '', text, flags=re.DOTALL)
                    content = text.encode('utf-8')
                zout.writestr(item, content)
    return fixed_path


def get_header_col_index(header_row: list, target_keywords: dict) -> dict:
    """根据关键词匹配列索引"""
    result = {}
    for key, keywords in target_keywords.items():
        for i, h in enumerate(header_row):
            hstr = str(h or "").strip()
            for kw in keywords:
                if kw in hstr:
                    result[key] = i
                    break
            if key in result:
                break
    return result


def compute_stats(values: list) -> list:
    """对一组值进行统计：返回 [(value, count, ratio), ...] 按 count 降序"""
    total = len(values)
    if total == 0:
        return []
    counter = Counter(values)
    sorted_items = sorted(counter.items(), key=lambda x: -x[1])
    return [(v, c, round(c / total, 4)) for v, c in sorted_items]


def find_all_villages(bd_ws, sheet1_ws):
    """查找所有村的多维汇总数据。

    返回 list[dict]:
      village_code, village_name,
      tasks: [{ task_name, total_area, total_parcels, dims: {dim_key: {label, stats}} }]
      village_dims: {dim_key: {label, stats, total_parcels}}
      village_total_parcels
    """
    bd_rows = list(bd_ws.iter_rows(values_only=True))
    if not bd_rows:
        return []
    bd_header = bd_rows[0]

    dim_keywords = {}
    for key, label, col_idx in DIMENSIONS:
        dim_keywords[key] = [label]

    col_map = get_header_col_index(bd_header, {
        "village_code": ["村代码"],
        "task_name": ["任务名称"],
        **dim_keywords,
    })

    vc_col = col_map.get("village_code")
    task_col = col_map.get("task_name")

    for key, label, _ in DIMENSIONS:
        if key not in col_map:
            print(f"警告: 村数据 sheet 找不到 '{label}' 列，跳过该维度")

    village_raw = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    village_order = []
    task_order = {}

    for r in bd_rows[1:]:
        vals = list(r)
        vc = str(vals[vc_col] or "").strip() if vc_col is not None and vc_col < len(vals) else ""
        tn = str(vals[task_col] or "").strip() if task_col is not None and task_col < len(vals) else ""
        if not tn or tn in ("None", "-", ""):
            continue
        if not vc or vc in ("None", "-", ""):
            continue
        if vc not in village_order:
            village_order.append(vc)
            task_order[vc] = []
        if tn not in village_raw[vc]:
            task_order[vc].append(tn)
        for key, label, col_idx in DIMENSIONS:
            col = col_map.get(key)
            val = str(vals[col] or "").strip() if col is not None and col < len(vals) else ""
            if val and val not in ("None", "-", "", "null"):
                village_raw[vc][tn][key].append(val)

    if not village_order:
        print("错误: 未找到任何有效的村数据")
        return []

    s1_rows = list(sheet1_ws.iter_rows(values_only=True))
    s1_header = s1_rows[0] if s1_rows else []
    s1_cols = get_header_col_index(s1_header, {
        "task_name": ["任务名称"],
        "total_area": ["地块总亩数"],
        "village_code": ["村代码"],
        "village_name": ["村名称"],
    })
    s1_task_col = s1_cols.get("task_name")
    s1_area_col = s1_cols.get("total_area")
    s1_vc_col = s1_cols.get("village_code")
    s1_vn_col = s1_cols.get("village_name")

    area_map = {}
    vname_map = {}

    if s1_task_col is not None and s1_area_col is not None:
        for r in s1_rows[1:]:
            tn = str(r[s1_task_col] or "").strip() if s1_task_col < len(r) else ""
            val = r[s1_area_col] if s1_area_col < len(r) else None
            vc = str(r[s1_vc_col] or "").strip() if s1_vc_col is not None and s1_vc_col < len(r) else ""
            vn = str(r[s1_vn_col] or "").strip() if s1_vn_col is not None and s1_vn_col < len(r) else ""
            if tn:
                try:
                    area_map[tn] = float(val) if val is not None else None
                except (ValueError, TypeError):
                    area_map[tn] = None
                if vc and vc not in vname_map:
                    vname_map[vc] = vn

    result = []
    for vc in village_order:
        v_total_parcels = 0
        v_all_values = defaultdict(list)
        task_list = []
        for tn in task_order[vc]:
            total_parcels = 0
            task_dims = {}
            for key, label, _ in DIMENSIONS:
                vals = village_raw[vc][tn].get(key, [])
                total_parcels = max(total_parcels, len(vals))
                task_dims[key] = {
                    "label": label,
                    "stats": compute_stats(vals),
                }
                v_all_values[key].extend(vals)
            task_list.append({
                "task_name": tn,
                "total_area": area_map.get(tn),
                "total_parcels": total_parcels,
                "dims": task_dims,
            })
            v_total_parcels += total_parcels

        village_dims = {}
        for key, label, _ in DIMENSIONS:
            village_dims[key] = {
                "label": label,
                "stats": compute_stats(v_all_values[key]),
                "total_parcels": len(v_all_values[key]),
            }
        result.append({
            "village_code": vc,
            "village_name": vname_map.get(vc, ""),
            "tasks": task_list,
            "village_dims": village_dims,
            "village_total_parcels": v_total_parcels,
        })
    return result


def read_comparison_data(wb) -> dict:
    """读取本地对比数据 sheet，返回 {village_short_name: {crop_type: area}}"""
    if "本地对比数据" not in wb.sheetnames:
        print("警告: 找不到'本地对比数据'sheet，跳过对比")
        return {}

    comp_ws = wb["本地对比数据"]
    rows = list(comp_ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}

    header = [str(c or "").strip() for c in rows[0]]
    # col0=村（社区）, rest=crop type names
    result = {}
    for r in rows[1:]:
        vname = str(r[0] or "").strip()
        if not vname:
            continue
        crop_map = {}
        for i in range(1, len(header)):
            if i < len(r) and r[i] is not None:
                try:
                    crop_map[header[i]] = float(r[i])
                except (ValueError, TypeError):
                    pass
        result[vname] = crop_map
    print(f"读取本地对比数据: {len(result)} 个村")
    return result


def extract_village_short_name(task_name: str) -> str:
    """从任务名 'AC乡/BD村村民委员会/001' 提取 'BD村'"""
    parts = task_name.split("/")
    if len(parts) >= 2:
        name = parts[1]
        # 村民委员会 → 村，社区居民委员会 → 社区
        name = name.replace("村民委员会", "村")
        name = name.replace("社区居民委员会", "社区")
        return name
    return task_name


def match_village_for_comparison(task_name: str, comp_villages: list) -> str:
    """在本地对比数据中找到匹配的村名"""
    short = extract_village_short_name(task_name)
    # 精确匹配
    if short in comp_villages:
        return short
    # 部分匹配
    for cv in comp_villages:
        if cv in short or short in cv:
            return cv
    return None


def write_dim_stats(ws, row, dim_data, total_area, col_start=1, level_label="", dim_label=""):
    """写入一个维度的统计数据（表头+明细行+小计），返回下一行号"""
    stats = dim_data["stats"]
    if not stats:
        return row

    if dim_label:
        ws.cell(row=row, column=col_start, value=dim_label)
        ws.cell(row=row, column=col_start).font = Font(bold=True, italic=True)
        for c in range(col_start, col_start + 4):
            ws.cell(row=row, column=c).fill = DIM_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    headers = [
        level_label or dim_data.get("label", "分类"),
        "地块数", "占比", "面积(亩)",
    ]
    for ci, h in enumerate(headers, col_start):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    row += 1

    total_cnt = 0
    for val, cnt, ratio in stats:
        ws.cell(row=row, column=col_start, value=val).border = THIN_BORDER
        ws.cell(row=row, column=col_start + 1, value=cnt).border = THIN_BORDER

        cell_r = ws.cell(row=row, column=col_start + 2, value=ratio)
        cell_r.border = THIN_BORDER
        cell_r.number_format = '0.0000'

        if isinstance(total_area, (int, float)) and total_area > 0:
            calc_area = round(ratio * total_area, 4)
            cell_a = ws.cell(row=row, column=col_start + 3, value=calc_area)
            cell_a.number_format = '#,##0.0000'
            cell_a.fill = AREA_FILL
        else:
            ws.cell(row=row, column=col_start + 3, value="N/A")
        ws.cell(row=row, column=col_start + 3).border = THIN_BORDER
        total_cnt += cnt
        row += 1

    # 小计行
    ws.cell(row=row, column=col_start, value="小计").border = THIN_BORDER
    ws.cell(row=row, column=col_start).font = Font(bold=True)
    ws.cell(row=row, column=col_start + 1, value=total_cnt).border = THIN_BORDER
    ws.cell(row=row, column=col_start + 2, value=1).border = THIN_BORDER
    if isinstance(total_area, (int, float)) and total_area > 0:
        cell_ta = ws.cell(row=row, column=col_start + 3, value=round(total_area, 4))
        cell_ta.number_format = '#,##0.0000'
    else:
        ws.cell(row=row, column=col_start + 3, value="N/A")
    ws.cell(row=row, column=col_start + 3).border = THIN_BORDER
    for c in range(col_start, col_start + 4):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
    row += 1
    return row


def write_village_sheet(ws, v, comparison_data):
    """写入单个村的完整分析 sheet（A-D 任务明细 + F-I 全村合计 + K-N 对比）"""
    V_COL, C_COL = 6, 11
    v_total_area = sum(t["total_area"] for t in v["tasks"] if isinstance(t["total_area"], (int, float)))

    # ── Left: task detail (A-D) ──
    row = 1
    title = f"村代码: {v['village_code']}  {v['village_name']}  ({len(v['tasks'])}个任务)"
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=13)
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = VILLAGE_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

    for t in v["tasks"]:
        a_str = f"{t['total_area']:.4f}" if isinstance(t['total_area'], (int, float)) else "N/A"
        ws.cell(row=row, column=1, value=f"任务: {t['task_name']}  |  地块数: {t['total_parcels']}  |  地块总亩数: {a_str}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = TASK_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
        for key, label, _ in DIMENSIONS:
            row = write_dim_stats(ws, row, t["dims"].get(key, {"stats": []}),
                                  t["total_area"], dim_label=label)
        row += 1

    # ── Right: village summary + comparison (F+, K+) ──
    row = 1

    # Titles
    ws.cell(row=row, column=V_COL, value=f"【全村合计】{v['village_name']}").font = Font(bold=True, size=12, color="C00000")
    for c in range(V_COL, V_COL + 4):
        ws.cell(row=row, column=c).fill = VILLAGE_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.cell(row=row, column=C_COL, value="【上报数据对比】").font = Font(bold=True, size=12, color="C00000")
    for c in range(C_COL, C_COL + 4):
        ws.cell(row=row, column=c).fill = VILLAGE_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

    for key, label, _ in DIMENSIONS:
        stats = v["village_dims"].get(key, {}).get("stats", [])
        if not stats:
            continue

        # Dimension label
        ws.cell(row=row, column=V_COL, value=f"【{label}】").font = Font(bold=True, italic=True)
        for c in range(V_COL, V_COL + 4):
            ws.cell(row=row, column=c).fill = DIM_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

        # Headers
        for ci, h in enumerate(["分类", "地块数", "占比", "面积(亩)"], V_COL):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER
        for ci, h in enumerate(["分类", "普查填报面积", "统计上报面积", "差值"], C_COL):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER
        row += 1

        matched_vname = match_village_for_comparison(v["tasks"][0]["task_name"], list(comparison_data.keys())) if comparison_data else None

        total_cnt = 0
        total_reported = 0
        for val, cnt, ratio in stats:
            calc_area = round(ratio * v_total_area, 4)
            total_cnt += cnt

            # Village summary
            for c, vv in [(V_COL, val), (V_COL + 1, cnt), (V_COL + 2, ratio), (V_COL + 3, calc_area)]:
                cell = ws.cell(row=row, column=c, value=vv)
                cell.border = THIN_BORDER
                if c == V_COL + 2: cell.number_format = '0.0000'
                if c == V_COL + 3: cell.number_format = '#,##0.0000'; cell.fill = AREA_FILL

            # Comparison
            ws.cell(row=row, column=C_COL, value=val).border = THIN_BORDER
            cell_ca = ws.cell(row=row, column=C_COL + 1, value=calc_area)
            cell_ca.number_format = '#,##0.0000'; cell_ca.border = THIN_BORDER

            mapped_name = COMPARISON_MAP.get(val)
            reported = None
            if matched_vname and mapped_name and mapped_name in comparison_data.get(matched_vname, {}):
                reported = comparison_data[matched_vname][mapped_name]

            if reported is not None:
                diff = round(calc_area - reported, 4)
                total_reported += reported

                cell_rep = ws.cell(row=row, column=C_COL + 2, value=round(reported, 4))
                cell_rep.number_format = '#,##0.0000'; cell_rep.border = THIN_BORDER

                cell_diff = ws.cell(row=row, column=C_COL + 3, value=diff)
                cell_diff.number_format = '#,##0.0000'; cell_diff.border = THIN_BORDER
                cell_diff.fill = POS_FILL if diff >= 0 else NEG_FILL
            else:
                for c in range(C_COL + 2, C_COL + 4):
                    ws.cell(row=row, column=c, value="-").border = THIN_BORDER
            row += 1

        # Totals
        ws.cell(row=row, column=V_COL, value=f"{label}合计").border = THIN_BORDER
        ws.cell(row=row, column=V_COL).font = Font(bold=True)
        ws.cell(row=row, column=V_COL + 1, value=total_cnt).border = THIN_BORDER
        ws.cell(row=row, column=V_COL + 2, value=1).border = THIN_BORDER
        cell_vta = ws.cell(row=row, column=V_COL + 3, value=round(v_total_area, 4))
        cell_vta.number_format = '#,##0.0000'; cell_vta.border = THIN_BORDER
        for c in range(V_COL, V_COL + 4):
            ws.cell(row=row, column=c).fill = TOTAL_FILL

        ws.cell(row=row, column=C_COL, value="合计").border = THIN_BORDER
        ws.cell(row=row, column=C_COL).font = Font(bold=True)
        cell_tot = ws.cell(row=row, column=C_COL + 1, value=round(v_total_area, 4))
        cell_tot.number_format = '#,##0.0000'; cell_tot.border = THIN_BORDER
        cell_rep_tot = ws.cell(row=row, column=C_COL + 2, value=round(total_reported, 4))
        cell_rep_tot.number_format = '#,##0.0000'; cell_rep_tot.border = THIN_BORDER
        diff_total = round(v_total_area - total_reported, 4)
        cell_diff_tot = ws.cell(row=row, column=C_COL + 3, value=diff_total)
        cell_diff_tot.number_format = '#,##0.0000'; cell_diff_tot.border = THIN_BORDER
        cell_diff_tot.fill = POS_FILL if diff_total >= 0 else NEG_FILL
        for c in range(C_COL + 1, C_COL + 4):
            ws.cell(row=row, column=c).fill = TOTAL_FILL
        row += 1

    # Column widths
    for col, w in {1: 52, 2: 16, 3: 16, 4: 16,
                   V_COL: 20, V_COL+1: 12, V_COL+2: 12, V_COL+3: 14,
                   C_COL: 18, C_COL+1: 14, C_COL+2: 14, C_COL+3: 14}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "B2"


def write_village_summary_sheet(ws, villages: list):
    """写入按村代码汇总的简洁报表"""
    headers = ["村代码", "村名称", "任务数", "总地块数", "地块总亩数"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for ri, v in enumerate(villages, 2):
        v_total_area = sum(t["total_area"] for t in v["tasks"] if isinstance(t["total_area"], (int, float)))
        ws.cell(row=ri, column=1, value=v["village_code"]).border = THIN_BORDER
        ws.cell(row=ri, column=2, value=v["village_name"]).border = THIN_BORDER
        ws.cell(row=ri, column=3, value=len(v["tasks"])).border = THIN_BORDER
        ws.cell(row=ri, column=4, value=v["village_total_parcels"]).border = THIN_BORDER
        cell = ws.cell(row=ri, column=5, value=round(v_total_area, 4))
        cell.number_format = '#,##0.0000'
        cell.border = THIN_BORDER

    col_widths = {1: 16, 2: 30, 3: 10, 4: 12, 5: 16}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def write_note_sheet(ws, total_villages: int):
    """写入使用说明"""
    dims_str = " / ".join([label for _, label, _ in DIMENSIONS])
    notes = [
        ["地块核查数据分析工具 - 使用说明"],
        [""],
        [f"共处理 {total_villages} 个村/社区"],
        [""],
        ["分析维度:"],
        [f"  {dims_str}"],
        [""],
        ["版面布局:"],
        ["  A-D列: 各任务明细数据（按村/任务分组）"],
        ["  F-I列: 全村合计（各维度聚合 + 面积计算）"],
        ["  K-N列: 上报数据对比（普查面积 vs 本地对比数据）"],
        [""],
        ["Step 1 - 按任务名称分组统计"],
        ["  从村数据 sheet 按任务名称分组，统计各地块属性的地块数和占比"],
        [""],
        ["Step 2 - 按村代码聚合"],
        ["  同一村代码下的所有任务合并统计，计算全村各地的块数和占比"],
        [""],
        ["Step 3 - 面积计算"],
        ["  将 Step 1/2 的占比 x Sheet1 中的地块总亩数，得到各属性的估算面积"],
        [""],
        ["对比说明:"],
        ["  上报面积从「本地对比数据」sheet 按村名+作物名匹配获取"],
        ["  差值 = 普查面积 - 上报面积"],
        [""],
        ["注意"],
        ["  一个村可能有多个任务，每个任务独立统计"],
        ["  未播种/空值不计入统计"],
    ]
    for i, row in enumerate(notes, 1):
        cell = ws.cell(row=i, column=1, value=row[0])
        if i == 1:
            cell.font = Font(bold=True, size=14)


def main():
    if len(sys.argv) < 2:
        print("用法: python land_analysis.py <Excel文件>")
        print("示例: python land_analysis.py 地块核查导出数据到村20260515.xlsx")
        sys.exit(1)

    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"错误: 文件不存在 {excel_path}")
        sys.exit(1)

    print(f"读取: {excel_path}")

    fixed_path = fix_excel_autofilter(excel_path)
    wb = load_workbook(fixed_path, data_only=True)
    names = wb.sheetnames
    print(f"工作表: {names}")

    if len(names) < 2:
        print("错误: Excel 至少需要 2 个 sheet（Sheet1 + 村数据）")
        wb.close()
        sys.exit(1)

    sheet1_ws = wb[names[0]]
    bd_ws = wb[names[1]]
    print(f"使用 Sheet1: {names[0]}")
    print(f"使用 村数据: {names[1]}")

    villages = find_all_villages(bd_ws, sheet1_ws)
    if not villages:
        print("错误: 未找到有效数据")
        wb.close()
        sys.exit(1)

    # ── 读取本地对比数据 ──
    comparison_data = read_comparison_data(wb)

    # ── 控制台输出 ──
    print(f"\n共找到 {len(villages)} 个村/社区代码:")
    for v in villages:
        v_total_area = sum(t["total_area"] for t in v["tasks"] if isinstance(t["total_area"], (int, float)))
        print(f"\n{'='*60}")
        print(f"【{v['village_code']}】{v['village_name']}  ({len(v['tasks'])}个任务, {v['village_total_parcels']}地块, {v_total_area:.4f}亩)")

        for t in v["tasks"]:
            a_s = f"{t['total_area']:.4f}亩" if isinstance(t['total_area'], (int, float)) else "N/A"
            print(f"\n  {t['task_name']}  ({t['total_parcels']}地块, {a_s})")
            for key, label, _ in DIMENSIONS:
                st = t["dims"].get(key, {}).get("stats", [])
                if st:
                    print(f"    [{label}]")
                    for val, cnt, ratio in st:
                        ca = round(ratio * t['total_area'], 4) if isinstance(t['total_area'], (int, float)) else 0
                        print(f"      {val}: {cnt}块, {ratio*100:.2f}%, {ca:.4f}亩")

        # 对比输出
        if comparison_data:
            vn = match_village_for_comparison(v["tasks"][0]["task_name"], list(comparison_data.keys()))
            if vn:
                print(f"\n  -- 上报数据对比 ({vn}) --")
                for key, label, _ in DIMENSIONS:
                    st = v["village_dims"].get(key, {}).get("stats", [])
                    for val, cnt, ratio in st:
                        ca = round(ratio * v_total_area, 4)
                        mn = COMPARISON_MAP.get(val)
                        if mn and mn in comparison_data.get(vn, {}):
                            rp = comparison_data[vn][mn]
                            diff = ca - rp
                            dp = diff / rp * 100 if rp else 0
                            print(f"    {val}: 计算{ca:.2f} / 上报{rp:.2f} = 差值{diff:+.2f} ({dp:+.2f}%)")

    # ── 写入新文件 ──
    out_wb = Workbook()
    first = True

    for v in villages:
        # 每个村生成独立 sheet，sheet 名取村简称（如 BD村、健康村）
        vname = extract_village_short_name(v["tasks"][0]["task_name"]) if v["tasks"] else v["village_code"]
        # 限制 sheet 名长度
        vname = vname[:31]
        if first:
            ws = out_wb.active
            ws.title = vname
            first = False
        else:
            ws = out_wb.create_sheet(vname)
        write_village_sheet(ws, v, comparison_data)
        print(f"  生成 sheet: {vname}")

    summary_ws = out_wb.create_sheet("村汇总")
    write_village_summary_sheet(summary_ws, villages)

    note_ws = out_wb.create_sheet("说明")
    write_note_sheet(note_ws, len(villages))

    if len(names) >= 3:
        orig = wb[names[2]]
        orows = list(orig.iter_rows(values_only=True))
        if len(orows) > 1:
            ref_ws = out_wb.create_sheet("原始参考")
            for ri, r in enumerate(orows, 1):
                for ci, v in enumerate(r, 1):
                    ref_ws.cell(row=ri, column=ci, value=v)

    base = excel_path.replace(".xlsx", "")
    out_path = f"{base}_分析结果.xlsx"
    if os.path.exists(out_path):
        out_path = f"{base}_分析结果_{int(time.time())}.xlsx"
    out_wb.save(out_path)
    print(f"\n结果已保存: {out_path}")
    wb.close()
    out_wb.close()
    if os.path.exists(fixed_path):
        for _ in range(5):
            try:
                os.remove(fixed_path)
                break
            except PermissionError:
                time.sleep(0.5)


if __name__ == "__main__":
    main()
