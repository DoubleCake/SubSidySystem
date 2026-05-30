"""
村联系人管理模块
- 村两委干部/工作人员信息 CRUD
- 农业负责人指定/切换
- Excel 导入/导出
"""
from fastapi import APIRouter, Depends, Query, UploadFile
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional
from database import get_db
from models import VillageContact, Village
from openpyxl import Workbook, load_workbook
from io import BytesIO
from datetime import datetime

router = APIRouter(prefix="/api/village-contacts", tags=["village-contacts"])

POSITIONS = ["书记", "副书记", "副主任", "文书", "其他"]


def _contact_out(c):
    return {
        "id": c.id, "village_id": c.village_id,
        "name": c.name, "phone": c.phone or "",
        "position": c.position or "", "is_agri_lead": bool(c.is_agri_lead),
        "sort_order": c.sort_order, "remark": c.remark or "",
        "created_at": c.created_at.isoformat() if c.created_at else "",
        "updated_at": c.updated_at.isoformat() if c.updated_at else "",
    }


@router.get("")
def list_contacts(
    village_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """获取联系人列表，可按村筛选"""
    q = db.query(VillageContact)
    if village_id:
        q = q.filter(VillageContact.village_id == village_id)
    contacts = q.order_by(VillageContact.sort_order.asc(), VillageContact.id.asc()).all()
    villages = {v.id: v.village_name for v in db.query(Village).all()}
    return {
        "items": [_contact_out(c) for c in contacts],
        "villages": [{"id": k, "village_name": v} for k, v in villages.items()],
    }


@router.post("")
def create_contact(payload: dict, db: Session = Depends(get_db)):
    """新增联系人"""
    c = VillageContact(
        village_id=payload["village_id"],
        name=payload["name"],
        phone=payload.get("phone", ""),
        position=payload.get("position", ""),
        is_agri_lead=payload.get("is_agri_lead", False),
        sort_order=payload.get("sort_order", 0),
        remark=payload.get("remark", ""),
    )
    # 如果设为农业负责人，先清除同村其他负责人的标记
    if c.is_agri_lead:
        db.query(VillageContact).filter(
            VillageContact.village_id == c.village_id,
            VillageContact.is_agri_lead == True
        ).update({VillageContact.is_agri_lead: False})
    db.add(c)
    db.commit()
    db.refresh(c)
    return _contact_out(c)


@router.put("/{contact_id}")
def update_contact(contact_id: int, payload: dict, db: Session = Depends(get_db)):
    """更新联系人信息"""
    c = db.query(VillageContact).filter(VillageContact.id == contact_id).first()
    if not c:
        return {"error": "联系人不存在"}
    for field in ["name", "phone", "position", "sort_order", "remark"]:
        if field in payload:
            setattr(c, field, payload[field])
    if "is_agri_lead" in payload:
        new_lead = bool(payload["is_agri_lead"])
        if new_lead and not c.is_agri_lead:
            db.query(VillageContact).filter(
                VillageContact.village_id == c.village_id,
                VillageContact.is_agri_lead == True
            ).update({VillageContact.is_agri_lead: False})
        c.is_agri_lead = new_lead
    db.commit()
    db.refresh(c)
    return _contact_out(c)


@router.delete("/{contact_id}")
def delete_contact(contact_id: int, db: Session = Depends(get_db)):
    """删除联系人"""
    c = db.query(VillageContact).filter(VillageContact.id == contact_id).first()
    if not c:
        return {"error": "联系人不存在"}
    db.delete(c)
    db.commit()
    return {"ok": True}


@router.post("/set-lead/{contact_id}")
def set_agri_lead(contact_id: int, db: Session = Depends(get_db)):
    """将指定联系人设为农业负责人（同一村只允许一个）"""
    c = db.query(VillageContact).filter(VillageContact.id == contact_id).first()
    if not c:
        return {"error": "联系人不存在"}
    db.query(VillageContact).filter(
        VillageContact.village_id == c.village_id,
        VillageContact.is_agri_lead == True
    ).update({VillageContact.is_agri_lead: False})
    c.is_agri_lead = True
    db.commit()
    return {"ok": True, "village_id": c.village_id, "name": c.name}


@router.get("/template")
def download_template(db: Session = Depends(get_db)):
    """下载联系人导入模板"""
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "联系人导入模板"
    ws.append(["村名", "姓名", "电话", "职务", "农业负责人", "备注"])
    # 示例数据
    villages = db.query(Village).order_by(Village.village_name).all()
    for v in villages[:3]:
        ws.append([v.village_name, "张三", "13800001111", "书记", "是", ""])
        ws.append([v.village_name, "李四", "13800002222", "副书记", "", ""])
    for col, w in zip("ABCDEF", [14, 10, 16, 12, 12, 20]):
        ws.column_dimensions[col].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=village_contacts_template.xlsx"}
    )


@router.post("/import")
async def import_contacts(
    file: UploadFile,
    village_id: int = Query(None),
    overwrite: bool = Query(False, description="是否覆盖已有同名联系人"),
    db: Session = Depends(get_db),
):
    """Excel 导入联系人：列顺序 村名/姓名/电话/职务/农业负责人/备注"""
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active

    # 解析表头做智能映射
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_map: dict[str, int] = {}
    for i, h in enumerate(header_row):
        if h is None: continue
        h = str(h).strip()
        if "村" in h: col_map["village"] = i
        elif "姓名" in h or "名称" in h: col_map["name"] = i
        elif "电话" in h or "手机" in h: col_map["phone"] = i
        elif "职务" in h or "职位" in h: col_map["position"] = i
        elif "负责" in h or "农业" in h: col_map["lead"] = i
        elif "备注" in h or "说明" in h: col_map["remark"] = i

    # 默认按列顺序
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0; updated = 0; errors = []

    # 预加载村名映射
    village_map = {v.village_name: v.id for v in db.query(Village).all()}

    for i, row in enumerate(rows, start=2):
        def cell(key: str, fallback_idx: int):
            idx = col_map.get(key, fallback_idx)
            return str(row[idx] or "").strip() if idx < len(row) else ""

        name = cell("name", 1)
        phone = cell("phone", 2)
        position = cell("position", 3)
        is_agri = cell("lead", 4)
        remark = cell("remark", 5)
        vname = cell("village", 0)

        if not name:
            errors.append(f"第{i}行：缺少姓名"); continue

        # 解析 village_id
        vid = village_id
        if not vid and vname:
            vid = village_map.get(vname)
            if not vid:
                errors.append(f"第{i}行：未找到村「{vname}」"); continue
        if not vid:
            errors.append(f"第{i}行：缺少村名且未指定村庄"); continue

        is_lead = is_agri in ("是", "yes", "Yes", "YES", "1")
        existing = None
        if overwrite:
            existing = db.query(VillageContact).filter(
                VillageContact.village_id == vid,
                VillageContact.name == name
            ).first()

        if existing:
            existing.phone = phone; existing.position = position; existing.remark = remark
            if is_lead and not existing.is_agri_lead:
                db.query(VillageContact).filter(
                    VillageContact.village_id == vid,
                    VillageContact.is_agri_lead == True
                ).update({VillageContact.is_agri_lead: False})
                existing.is_agri_lead = True
            updated += 1
        else:
            c = VillageContact(
                village_id=vid, name=name, phone=phone,
                position=position, is_agri_lead=is_lead,
                sort_order=created + updated, remark=remark,
            )
            if is_lead:
                db.query(VillageContact).filter(
                    VillageContact.village_id == vid,
                    VillageContact.is_agri_lead == True
                ).update({VillageContact.is_agri_lead: False})
            db.add(c)
            created += 1
    db.commit()
    return {"created": created, "updated": updated, "errors": errors}


@router.get("/export")
def export_contacts(
    village_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """Excel 导出联系人"""
    from fastapi.responses import StreamingResponse

    q = db.query(VillageContact)
    if village_id:
        q = q.filter(VillageContact.village_id == village_id)
    contacts = q.order_by(VillageContact.sort_order.asc(), VillageContact.id.asc()).all()
    villages = {v.id: v.village_name for v in db.query(Village).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "村联系人"
    ws.append(["村名", "姓名", "电话", "职务", "农业负责人", "备注"])
    for c in contacts:
        ws.append([
            villages.get(c.village_id, ""), c.name, c.phone or "",
            c.position or "", "是" if c.is_agri_lead else "",
            c.remark or ""
        ])
    # 列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=village_contacts_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )
