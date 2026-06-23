"""
数据备份与恢复路由
- GET  /api/backup/download      → 下载完整数据库文件（.db）
- GET  /api/backup/export-excel  → 导出全量 Excel（多 sheet）
- POST /api/backup/restore       → 上传 .db 文件恢复（谨慎）
- GET  /api/backup/info          → 当前数据库统计信息
- POST /api/backup/auto          → 创建带时间戳的本地备份文件
"""

import os, shutil, json
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
import io

from database import get_db, engine, DATABASE_URL, DB_PATH
from utils import format_group_no

router = APIRouter(prefix="/api/backup", tags=["备份管理"])

# 备份目录
BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backups")


def ensure_backup_dir():
    os.makedirs(BACKUP_DIR, exist_ok=True)


# ─── 数据库信息 ───
@router.get("/info")
def backup_info(db: Session = Depends(get_db)):
    tables = {
        "farmer_profile":      "farmer_profile",
        "family_household":    "family_household",
        "village_group":       "village_group",
        "subsidy_type":        "subsidy_type",
        "subsidy_application": "subsidy_application",
    }
    counts = {}
    for key, tbl in tables.items():
        try:
            counts[key] = db.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
        except Exception:
            counts[key] = 0

    db_size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0

    # 列出已有备份
    ensure_backup_dir()
    backups = []
    for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
        if f.endswith(".db"):
            fpath = os.path.join(BACKUP_DIR, f)
            backups.append({
                "filename": f,
                "size_kb": round(os.path.getsize(fpath) / 1024, 1),
                "created": datetime.fromtimestamp(os.path.getmtime(fpath)).strftime("%Y-%m-%d %H:%M:%S"),
            })

    return {
        "db_path":     DB_PATH,
        "db_size_kb":  round(db_size / 1024, 1),
        "db_size_mb":  round(db_size / 1024 / 1024, 2),
        "record_counts": counts,
        "total_records": sum(counts.values()),
        "backups": backups[:10],  # 最近10个
        "backup_dir": BACKUP_DIR,
    }


# ─── 下载数据库文件 ───
@router.get("/download")
def download_db():
    if not os.path.exists(DB_PATH):
        raise HTTPException(status_code=404, detail="数据库文件不存在")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return FileResponse(
        path=DB_PATH,
        filename=f"subsidy_backup_{ts}.db",
        media_type="application/octet-stream",
    )


# ─── 创建本地时间戳备份 ───
@router.post("/auto")
def create_local_backup():
    if not os.path.exists(DB_PATH):
        raise HTTPException(status_code=404, detail="数据库文件不存在")
    ensure_backup_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(BACKUP_DIR, f"subsidy_{ts}.db")
    shutil.copy2(DB_PATH, dest)
    size_kb = round(os.path.getsize(dest) / 1024, 1)
    return {"message": f"备份成功", "filename": f"subsidy_{ts}.db",
            "path": dest, "size_kb": size_kb}


# ─── 导出全量 Excel（多 sheet）───
@router.get("/export-excel")
def export_excel(db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="请先安装 openpyxl：pip install openpyxl")

    wb = openpyxl.Workbook()

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2D6A4F")
    center_align = Alignment(horizontal="center")

    def write_sheet(ws, headers: list, rows: list):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Sheet 1: 农户档案
    ws1 = wb.active; ws1.title = "农户档案"
    rows1 = db.execute(text("""
        SELECT fp.id, fp.real_name,
               CASE fp.gender WHEN 1 THEN '男' ELSE '女' END,
               fp.id_card, fp.phone, fp.bank_card, fp.bank_name,
               CASE WHEN hh.head_farmer_id = fp.id THEN '户主' ELSE '成员' END,
               fp.relation,
               CASE fp.farmer_status WHEN 1 THEN '在册' WHEN 2 THEN '注销'
                 WHEN 3 THEN '迁出' WHEN 4 THEN '死亡' ELSE '未知' END,
               COALESCE(v.village_name, '') AS village_name,
               COALESCE(hh.group_no, 1) AS group_no,
               hh.contract_area, hh.household_code,
               fp.remark, fp.created_at
        FROM farmer_profile fp
        LEFT JOIN family_household hh ON fp.household_id = hh.id
        LEFT JOIN village v ON v.id = hh.village_id
        ORDER BY v.village_name, hh.group_no, fp.id
    """)).fetchall()

    def _fmt_row1(r):
        vn = r.village_name or ''
        gn = format_group_no(r.group_no) if r.group_no else '一组'
        row = list(r)
        row[10] = f"{vn}{gn}"
        del row[11]  # 删除单独的 group_no 列，避免列数与表头不符
        return row
    write_sheet(ws1,
        ["ID","姓名","性别","身份证号","手机号","银行卡号","开户行",
         "角色","与户主关系","状态","所在村组","土地面积(亩)","家庭户编码","备注","录入时间"],
        [_fmt_row1(r) for r in rows1])

    # Sheet 2: 家庭户
    ws2 = wb.create_sheet("家庭户")
    rows2 = db.execute(text("""
        SELECT hh.id, hh.household_code, hh.household_name,
               COALESCE(v.village_name, '') AS village_name,
               COALESCE(hh.group_no, 1) AS group_no,
               hh.contract_area,
               fp.real_name AS head_name,
               CASE hh.status WHEN 1 THEN '正常' WHEN 2 THEN '注销' ELSE '其他' END,
               hh.address
        FROM family_household hh
        LEFT JOIN village v ON v.id = hh.village_id
        LEFT JOIN farmer_profile  fp ON hh.head_farmer_id = fp.id
        ORDER BY v.village_name, hh.group_no, hh.id
    """)).fetchall()

    def _fmt_row2(r):
        vn = r.village_name or ''
        gn = format_group_no(r.group_no) if r.group_no else '一组'
        row = list(r)
        row[3] = f"{vn}{gn}"
        del row[4]  # 删除单独的 group_no 列
        return row
    write_sheet(ws2,
        ["ID","户编码","户名","所在村组","承包面积(亩)","户主姓名","状态","地址"],
        [_fmt_row2(r) for r in rows2])

    # Sheet 3: 补贴申请记录
    ws3 = wb.create_sheet("补贴申请记录")
    rows3 = db.execute(text("""
        SELECT sa.id, fp.real_name, fp.id_card,
               COALESCE(v.village_name, '') AS village_name,
               COALESCE(hh.group_no, 1) AS group_no,
               st.subsidy_name, st.subsidy_year,
               CASE st.calc_mode WHEN 'per_mu' THEN '按亩' ELSE '固定' END,
               sa.apply_area, sa.apply_amount, sa.actual_amount,
               CASE sa.pay_status WHEN 0 THEN '待发放' WHEN 1 THEN '部分发放'
                 WHEN 2 THEN '已发放' WHEN 3 THEN '驳回' ELSE '未知' END,
               sa.pay_date, sa.remark
        FROM subsidy_application sa
        LEFT JOIN farmer_profile  fp ON sa.farmer_id = fp.id
        LEFT JOIN subsidy_type    st ON sa.subsidy_type_id = st.id
        LEFT JOIN family_household hh ON fp.household_id = hh.id
        LEFT JOIN village v ON v.id = hh.village_id
        ORDER BY st.subsidy_year DESC, v.village_name, hh.group_no, fp.id
    """)).fetchall()

    def _fmt_row3(r):
        vn = r.village_name or ''
        gn = format_group_no(r.group_no) if r.group_no else '一组'
        row = list(r)
        row[3] = f"{vn}{gn}"
        del row[4]  # 删除单独的 group_no 列
        return row
    write_sheet(ws3,
        ["ID","姓名","身份证号","所在村组","补贴项目","年度","计算方式",
         "申请面积(亩)","申请金额","实发金额","发放状态","打款日期","备注"],
        [_fmt_row3(r) for r in rows3])

    # Sheet 4: 补贴项目
    ws4 = wb.create_sheet("补贴项目")
    rows4 = db.execute(text("""
        SELECT id, subsidy_name, subsidy_year,
               CASE calc_mode WHEN 'per_mu' THEN '按亩' ELSE '固定' END,
               standard_amount, standard_unit, fund_source, apply_deadline,
               CASE pay_status WHEN 0 THEN '未发放' WHEN 1 THEN '部分' WHEN 2 THEN '已完成' ELSE '未知' END,
               description
        FROM subsidy_type ORDER BY subsidy_year DESC, id
    """)).fetchall()
    write_sheet(ws4,
        ["ID","补贴名称","年度","计算方式","标准金额","单位","资金来源","截止日期","发放状态","说明"],
        [list(r) for r in rows4])

    # Sheet 5: 村组（从 village + family_household.group_no 聚合）
    ws5 = wb.create_sheet("村组配置")
    rows5 = db.execute(text("""
        SELECT DISTINCT v.id, v.village_name, hh.group_no
        FROM village v
        LEFT JOIN family_household hh ON hh.village_id = v.id
        ORDER BY v.village_name, hh.group_no
    """)).fetchall()

    def _fmt_row5(r):
        row = list(r)
        # 组号存的是整数，显示时转为"X组"
        gn = row[2] if r.group_no else 1
        row[2] = format_group_no(gn)
        return row
    write_sheet(ws5, ["ID","村名","组号"], [_fmt_row5(r) for r in rows5])

    # 写入内存并返回
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=subsidy_export_{ts}.xlsx"}
    )


# ─── 恢复数据库（上传 .db 文件）───
@router.post("/restore")
async def restore_db(file: UploadFile = File(...)):
    if not file.filename or not file.filename.endswith(".db"):
        raise HTTPException(status_code=400, detail="请上传 .db 格式的数据库文件")

    # 先备份当前数据库
    ensure_backup_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if os.path.exists(DB_PATH):
        shutil.copy2(DB_PATH, os.path.join(BACKUP_DIR, f"before_restore_{ts}.db"))

    content = await file.read()

    # 简单验证：SQLite 文件头
    if not content.startswith(b"SQLite format 3"):
        raise HTTPException(status_code=400, detail="文件不是有效的 SQLite 数据库")

    with open(DB_PATH, "wb") as f:
        f.write(content)

    return {
        "message": "数据库恢复成功，请重启后端服务使更改生效",
        "backup_created": f"before_restore_{ts}.db",
        "restored_size_kb": round(len(content) / 1024, 1),
    }


# ─── 下载指定备份文件 ───
@router.get("/backups/{filename}")
def download_backup(filename: str):
    if ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="非法文件名")
    path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="备份文件不存在")
    return FileResponse(path=path, filename=filename, media_type="application/octet-stream")


# ─── 删除备份文件 ───
@router.delete("/backups/{filename}")
def delete_backup(filename: str):
    if ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="非法文件名")
    path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="备份文件不存在")
    os.remove(path)
    return {"message": f"{filename} 已删除"}
