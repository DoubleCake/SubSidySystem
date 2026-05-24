"""
生成公式驱动的 Excel 分析工作簿
Sheet1=地块属性, Sheet2=种植属性, Sheet3=本地对比数据 (源数据)
Sheet4=查询分析 (输入村名，公式动态计算)
"""
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula

SRC = "地块核查导出数据到村20260515.xlsx"
OUT = "地块核查_公式分析表.xlsx"

# ── 样式 ──
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SEC_FONT = Font(bold=True, size=12, color="2E75B6")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
HDR_FILL = PatternFill("solid", fgColor="4472C4")
LBL_FONT = Font(bold=True, size=11)
INFO_FONT = Font(size=11)
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── 读取源数据 ──
src = load_workbook(SRC, data_only=True)

# ── 收集唯一村名（用于下拉列表） ──
src1 = src["地块属性"]
village_names = []
seen = set()
for r in range(2, src1.max_row + 1):
    name = src1.cell(r, 7).value
    if name and name not in seen:
        seen.add(name)
        village_names.append(name)

# ── 创建输出工作簿 ──
wb = Workbook()

# ================================================================
# Sheet1: 地块属性
# ================================================================
ws1 = wb.active
ws1.title = "地块属性"
src1_src = src["地块属性"]
for r_idx, row in enumerate(src1_src.iter_rows(min_row=1, max_row=src1_src.max_row, max_col=src1_src.max_column), 1):
    ws1.cell(r_idx, 1).value = None  # reset
    for c_idx, cell in enumerate(row, 1):
        ws1.cell(r_idx, c_idx, cell.value)
# Freeze header
ws1.freeze_panes = "A2"

# ================================================================
# Sheet2: 种植属性
# ================================================================
ws2 = wb.create_sheet("种植属性")
src2_src = src["种植属性"]
total_rows = src2_src.max_row
# Batch append is faster than cell-by-cell
for row in src2_src.iter_rows(min_row=1, max_row=total_rows, max_col=src2_src.max_column):
    ws2.append([c.value for c in row])
ws2.freeze_panes = "A2"

# ================================================================
# Sheet3: 本地对比数据
# ================================================================
ws3 = wb.create_sheet("本地对比数据")
src3_src = src["本地对比数据"]
for row in src3_src.iter_rows(min_row=1, max_row=src3_src.max_row, max_col=src3_src.max_column):
    ws3.append([c.value for c in row])
ws3.freeze_panes = "A2"

# ================================================================
# Sheet4: 查询分析（公式驱动）
# ================================================================
ws4 = wb.create_sheet("查询分析")
ws4.sheet_properties.tabColor = "2E75B6"

# Column widths
for col, w in {1: 6, 2: 28, 3: 16, 4: 16, 5: 16}.items():
    ws4.column_dimensions[get_column_letter(col)].width = w

# ── 辅助区域：Z1 = 作物名映射表 ──
# Mapping: our parcel crop name → Sheet3 column name
MAPPING = [
    ("旱地", "旱地", "G"),         # G = 农作物种植用地属性
    ("其他水田", "其他水田", "G"),
    ("水浇地", "水浇地", "G"),
    ("冬水田", "冬水田", "G"),
    ("小麦", "小麦", "H"),         # H = 夏收主要作物
    ("油菜", "油菜籽", "H"),
    ("马铃薯", "马铃薯", "H"),
    ("蔬菜", "蔬菜（含菜用瓜）", "H"),
    ("其他经济作物", "中草药材", "H"),
    ("玉米", "玉米", "M"),         # M = 秋收主要作物
    ("水稻", "稻谷", "M"),
    ("大豆", "大豆", "M"),
    ("甘薯", "甘薯", "M"),
    ("高粱", "高粱", "M"),
    ("花生", "花生", "M"),
]

# Map Sheet3 columns: we need the column letter for each crop type in Sheet3
# Sheet3 row 1 has crop names as headers in cols B-Q (col2-col17)
# Get crop→Sheet3_col mapping
sheet3_crop_col = {}
for c in range(2, src3_src.max_column + 1):
    h = src3_src.cell(1, c).value
    if h:
        sheet3_crop_col[h] = get_column_letter(c)

# Write mapping table to Z1:AA3 area (hidden helper zone)
# Z1=源列(Z=种植属性列), Z2=原名称, Z3=映射名
ws4["Z1"] = "源列"
ws4["AA1"] = "原名称"
ws4["AB1"] = "映射名"
ws4["AC1"] = "Sheet3列"
for i, (orig, mapped, src_col) in enumerate(MAPPING, 2):
    ws4.cell(i, 26, src_col)   # Z = source column letter
    ws4.cell(i, 27, orig)      # AA = original crop name
    ws4.cell(i, 28, mapped)    # AB = mapped crop name
    sc = sheet3_crop_col.get(mapped, "")
    ws4.cell(i, 29, sc)        # AC = Sheet3 column letter

# Hide helper columns
for c in ["Z", "AA", "AB", "AC", "AD"]:
    ws4.column_dimensions[c].hidden = True

# ── 标题行 ──
ws4.merge_cells("A1:D1")
ws4.cell(1, 1, "📋 地块核查数据查询分析").font = TITLE_FONT

# ── 输入区 ──
ws4.cell(2, 1, "选择村委会：").font = LBL_FONT
ws4.cell(2, 1).alignment = Alignment(horizontal="right")
ws4.cell(2, 2).font = INFO_FONT
ws4.cell(2, 2).border = THIN

# Data validation dropdown for village name
# Write unique names to AD column for the dropdown
for i, name in enumerate(village_names, 1):
    ws4.cell(i, 30, name)  # AD1, AD2, ...
dv = DataValidation(
    type="list",
    formula1=f"=$AD$1:$AD${len(village_names)}",
    allow_blank=True,
    showDropDown=False,
)
dv.error = "请选择有效的村委会名称"
dv.errorTitle = "无效选择"
dv.prompt = "请从下拉列表选择村委会"
dv.promptTitle = "选择村委会"
ws4.add_data_validation(dv)
dv.add("B2")

# B3: Village code
ws4.cell(3, 1, "村代码：").font = LBL_FONT
ws4.cell(3, 1).alignment = Alignment(horizontal="right")
ws4.cell(3, 2).font = INFO_FONT
ws4.cell(3, 2).border = THIN
ws4.cell(3, 2).value = '=IF(B2="","",XLOOKUP(B2,地块属性!G:G,地块属性!F:F))'

# B4: Total parcels for this village
ws4.cell(4, 1, "总地块数：").font = LBL_FONT
ws4.cell(4, 1).alignment = Alignment(horizontal="right")
ws4.cell(4, 2).font = INFO_FONT
ws4.cell(4, 2).border = THIN
ws4.cell(4, 2).value = '=IF(B3="","",COUNTIFS(种植属性!C:C,B3))'

# B5: Total area for this village
ws4.cell(5, 1, "地块总面积(亩)：").font = LBL_FONT
ws4.cell(5, 1).alignment = Alignment(horizontal="right")
ws4.cell(5, 2).font = INFO_FONT
ws4.cell(5, 2).border = THIN
ws4.cell(5, 2).value = '=IF(B3="","",SUMIFS(地块属性!M:M,地块属性!F:F,B3))'
ws4.cell(5, 2).number_format = '#,##0.0000'

# ── 一、任务包概况 ──
row = 7
ws4.cell(row, 1, "一、任务包概况").font = SEC_FONT
row = 8
headers = ["序号", "任务名称", "地块数", "地块总亩数"]
for ci, h in enumerate(headers, 1):
    cell = ws4.cell(row, ci, h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN

row = 9
# A9: =IF(B3="","",UNIQUE(FILTER(种植属性!D2:D5000,种植属性!C2:C5000=B3)))
ws4.cell(row, 1).value = '=IF(B3="","",SEQUENCE(ROWS(B9#)))'
# B9: unique task names
ws4.cell(row, 2).value = '=IF(B3="","",UNIQUE(FILTER(种植属性!D2:D5000,种植属性!C2:C5000=B3)))'
# C9: parcel count per task
ws4.cell(row, 3).value = '=IF(B3="","",COUNTIFS(种植属性!C:C,B3,种植属性!D:D,B9#))'
# D9: total area per task
ws4.cell(row, 4).value = '=IF(B3="","",SUMIFS(地块属性!M:M,地块属性!F:F,B3,地块属性!H:H,B9#))'
ws4.cell(row, 4).number_format = '#,##0.0000'

# ── 二、农作物种植用地属性 ──
row = 12
ws4.cell(row, 1, "二、农作物种植用地属性（全村汇总）").font = SEC_FONT
row = 13
for ci, h in enumerate(["分类", "地块数", "占比", "估算面积(亩)"], 1):
    cell = ws4.cell(row, ci, h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN

row = 14
# A14: unique crop types for this dimension
ws4.cell(row, 1).value = '=IF(B3="","",UNIQUE(FILTER(种植属性!G2:G5000,(种植属性!C2:C5000=B3)*(种植属性!G2:G5000<>""))))'
# B14: count per crop
ws4.cell(row, 2).value = '=IF(B3="","",COUNTIFS(种植属性!C:C,B3,种植属性!G:G,A14#))'
# C14: ratio
ws4.cell(row, 3).value = '=IF(B3="","",B14#/SUM(B14#))'
ws4.cell(row, 3).number_format = '0.00%'
# D14: estimated area
ws4.cell(row, 4).value = '=IF(B3="","",C14#*B5)'
ws4.cell(row, 4).number_format = '#,##0.0000'

# ── 三、2025年夏收主要作物 ──
row = 17
ws4.cell(row, 1, "三、2025年夏收主要作物（全村汇总）").font = SEC_FONT
row = 18
for ci, h in enumerate(["分类", "地块数", "占比", "估算面积(亩)"], 1):
    cell = ws4.cell(row, ci, h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN

row = 19
ws4.cell(row, 1).value = '=IF(B3="","",UNIQUE(FILTER(种植属性!H2:H5000,(种植属性!C2:C5000=B3)*(种植属性!H2:H5000<>""))))'
ws4.cell(row, 2).value = '=IF(B3="","",COUNTIFS(种植属性!C:C,B3,种植属性!H:H,A19#))'
ws4.cell(row, 3).value = '=IF(B3="","",B19#/SUM(B19#))'
ws4.cell(row, 3).number_format = '0.00%'
ws4.cell(row, 4).value = '=IF(B3="","",C19#*B5)'
ws4.cell(row, 4).number_format = '#,##0.0000'

# ── 四、2025年秋收主要作物 ──
row = 22
ws4.cell(row, 1, "四、2025年秋收主要作物（全村汇总）").font = SEC_FONT
row = 23
for ci, h in enumerate(["分类", "地块数", "占比", "估算面积(亩)"], 1):
    cell = ws4.cell(row, ci, h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN

row = 24
ws4.cell(row, 1).value = '=IF(B3="","",UNIQUE(FILTER(种植属性!M2:M5000,(种植属性!C2:C5000=B3)*(种植属性!M2:M5000<>""))))'
ws4.cell(row, 2).value = '=IF(B3="","",COUNTIFS(种植属性!C:C,B3,种植属性!M:M,A24#))'
ws4.cell(row, 3).value = '=IF(B3="","",B24#/SUM(B24#))'
ws4.cell(row, 3).number_format = '0.00%'
ws4.cell(row, 4).value = '=IF(B3="","",C24#*B5)'
ws4.cell(row, 4).number_format = '#,##0.0000'

# ── 五、上报数据对比 ──
row = 27
ws4.cell(row, 1, "五、上报数据对比").font = SEC_FONT
row = 28
for ci, h in enumerate(["分类", "普查估算面积", "统计上报面积", "差值"], 1):
    cell = ws4.cell(row, ci, h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN

# Comparison table: list crops from the mapping table
# For each crop, compute:
#   A: crop name from mapping (original)
#   B: estimated area (count × total_area / total_parcels) - need to search in the right dimension column
#   C: reported area from Sheet3
#   D: difference (B-C)

# Using a simple approach: list all mapped crops and compute
row = 29
# We'll use a fixed list of crops from the mapping (AA2:AA16)
# A29: =IF(B3="","",FILTER(AA2:AA16, AA2:AA16<>""))
# But this uses AA range which has our mapping data, including empty rows
# Let me use a fixed number of crops from the mapping

# Actually, let me take the crop list from the mapping table
# For each crop in mapping (AA2:AA16), count occurrences in Sheet2
# across the right dimension column (Z2)

# A29: get list of mapped crop names
# B29: count occurrences in the right column, then compute area
# C29: XLOOKUP from Sheet3
# D29: B-C

# Simplified: just enumerate the mapping table
# A29: =IF(B3="","",FILTER(AA2:AA16, AA2:AA16<>""))
ws4.cell(row, 1).value = '=IF(B3="","",FILTER(AA2:AA16,AA2:AA16<>""))'

# B29: count in the right source column, × (B5/B4) to get area
# Use INDIRECT to construct dynamic range: "种植属性!" & Z_column & ":" & Z_column
# Then COUNTIFS
# COUNTIFS(INDIRECT("种植属性!"&Z2&":"&Z2), B3&"", INDIRECT("种植属性!"&Z2&":"&Z2), A29)
# Wait, this doesn't work because INDIRECT with spill arrays is tricky.

# Simpler approach: For each crop, search in ALL 3 dimension columns
# and sum the counts, then compute area
# B29: = (COUNTIFS(种植属性!C:C,B3,种植属性!G:G,A29#) +
#         COUNTIFS(种植属性!C:C,B3,种植属性!H:H,A29#) +
#         COUNTIFS(种植属性!C:C,B3,种植属性!M:M,A29#)) * B5 / B4

ws4.cell(row, 2).value = '=IF(B3="","",(COUNTIFS(种植属性!C:C,B3,种植属性!G:G,A29#)+COUNTIFS(种植属性!C:C,B3,种植属性!H:H,A29#)+COUNTIFS(种植属性!C:C,B3,种植属性!M:M,A29#))*B5/MAX(B4,1))'
ws4.cell(row, 2).number_format = '#,##0.0000'

# C29: lookup reported area from Sheet3
# The crop name from mapping (AA2:AA16) needs to map to Sheet3 column
# XLOOKUP(A29, AB2:AB16, AC2:AC16) → get Sheet3 column letter
# Then INDIRECT("本地对比数据!" & col_letter & row) to get value

# Actually, let me use a helper: For each crop name in column A,
# look up the Sheet3 column via the mapping (AB→AC),
# then get that column value from Sheet3 row matching village short name

# First, get village short name for Sheet3 lookup
# B4 has the village full name like "BD村村民委员会", need "BD村"
# Let me add a cell with the short name
# Actually, let me just use XLOOKUP with the long name on Sheet3
# Sheet3!A:A has names like "BD村", so full name won't match
# Need to extract short name

# Let me simplify: write "BD村" style names directly from what we know
# E4: short name for Sheet3 lookup
# =LEFT(B2, FIND("村", B2)-1) & "村"  → "BD村民委员会" → "BD村"?
# Hmm, that doesn't work for all cases.

# Simplest: use the known village name from Sheet3 directly with XLOOKUP
# The village names in Sheet3 are: EF村, BD村
# The village names in 地块属性 are like "BD村村民委员会"
# We need to extract just the village part

# =LEFT(B2, FIND("村", B2)) → "BD村" for "BD村村民委员会"
# =LEFT(B2, FIND("村", B2)) → "BD村" ← correct!

ws4.cell(4, 4, "Sheet3村名").font = Font(size=9, color="808080")
ws4.cell(5, 4).value = '=IF(B2="","",LEFT(B2,FIND("村",B2)))'
ws4.cell(5, 4).font = INFO_FONT

# C29: XLOOKUP to find reported area
# XLOOKUP($D$5, 本地对比数据!A:A, INDIRECT("本地对比数据!"&XLOOKUP(A29,AB2:AB16,AC2:AC16)&":"&XLOOKUP(A29,AB2:AB16,AC2:AC16)))
# This is complex. Let me simplify: use XLOOKUP on Sheet3 directly
# For each crop in A29, the column in Sheet3 is found via: XLOOKUP(A29, AB2:AB16, AC2:AC16)
# Then INDIRECT("本地对比数据!" & col & ":" & col) to get whole column
# Then XLOOKUP to find the matching row

ws4.cell(row, 3).value = '=IF(B3="","",XLOOKUP($D$5,本地对比数据!A:A,XLOOKUP(A29#,AB2:AB16,AC2:AC16)))'
ws4.cell(row, 3).number_format = '#,##0.0000'

# D29: difference
ws4.cell(row, 4).value = '=IF(B3="","",B29#-C29#)'
ws4.cell(row, 4).number_format = '#,##0.0000'

# Conditional formatting for difference column (D29:D44)
# We'll use openpyxl's conditional formatting
from openpyxl.formatting.rule import CellIsRule
ws4.conditional_formatting.add(
    f"D29:D44",
    CellIsRule(operator="greaterThan", formula=["0"],
               fill=GREEN_FILL)
)
ws4.conditional_formatting.add(
    f"D29:D44",
    CellIsRule(operator="lessThan", formula=["0"],
               fill=RED_FILL)
)

# ── 冻结和打印 ──
ws4.freeze_panes = "A2"
ws4.sheet_view.showGridLines = True

# ── 保存 ──
wb.save(OUT)
print(f"已生成: {OUT}")
print(f"  Sheet1: 地块属性 ({src1_src.max_row} 行)")
print(f"  Sheet2: 种植属性 ({total_rows} 行)")
print(f"  Sheet3: 本地对比数据 ({src3_src.max_row} 行)")
print(f"  Sheet4: 查询分析 (公式驱动)")
print(f"  村名下拉列表: {len(village_names)} 个村")
