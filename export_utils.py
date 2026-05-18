"""
Excel 导出工具（使用 openpyxl）
提供更美观的样式和更好的格式控制
"""
from io import BytesIO
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from typing import Any, Dict, List, Optional


# ─── 样式定义 ───
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
BORDER_THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ALTERNATE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def auto_width(ws, max_width=50):
    """自动调整列宽"""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column].width = adjusted_width


def set_row_height(ws, header_height=25, data_height=40):
    """设置行高"""
    ws.row_dimensions[1].height = header_height
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = data_height


def apply_styles(ws, has_alternate_rows=True):
    """应用样式到整个工作表"""
    # 表头样式
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

    # 数据行样式
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_alternate = has_alternate_rows and (row_idx % 2 == 0)
        for cell in row:
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = BORDER_THIN
            if is_alternate:
                cell.fill = ALTERNATE_FILL

    auto_width(ws, max_width=50)
    set_row_height(ws)


def add_summary_sheet(wb: Workbook, summary: Dict[str, Any]):
    """添加汇总表"""
    ws = wb.create_sheet("汇总")

    # 准备数据
    data = [
        ["项目", "数值"],
        ["检查总行数", summary.get("total_rows", 0)],
        ["通过行数", summary.get("ok_rows", 0)],
        ["错误行数", summary.get("error_rows", 0)],
        ["通过率", f"{summary.get('pass_rate', 0)}%"],
        ["格式错误", summary.get("format_errors", 0)],
        ["村组不存在", summary.get("village_errors", 0)],
        ["重复身份证", summary.get("duplicate_errors", 0)],
        ["数据库已有申请记录", summary.get("db_duplicate_apps", 0)],
        ["性别不符", summary.get("gender_mismatch", 0)],
        ["错误库命中", summary.get("error_library_hits", 0)],
        ["面积异常", summary.get("area_anomalies", 0)],
        ["面积缺失", summary.get("area_missing", 0)],
        ["年龄异常", summary.get("age_anomaly", 0)],
        ["死亡农户", summary.get("deceased_farmers", 0)],
        ["受限身份农户", summary.get("restricted_farmers", 0)],
        ["同一家庭多成员申请", summary.get("household_duplicates", 0)],
        ["新增农户", summary.get("new_farmers", 0)],
        ["减少农户", summary.get("removed_farmers", 0)],
        ["字段变更", summary.get("changed_farmers", 0)],
    ]

    for row in data:
        ws.append(row)

    apply_styles(ws, has_alternate_rows=True)
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15


def add_sheet_from_data(wb: Workbook, sheet_name: str, headers: List[str], rows: List[Dict[str, Any]]):
    """从数据添加工作表"""
    if not rows:
        return

    ws = wb.create_sheet(sheet_name[:31])

    # 添加表头
    ws.append(headers)

    # 添加数据行
    for row in rows:
        row_data = []
        for h in headers:
            val = row.get(h, "")
            if isinstance(val, (list, tuple)):
                val = "；".join(str(x) for x in val)
            row_data.append(val)
        ws.append(row_data)

    apply_styles(ws, has_alternate_rows=True)


def export_precheck_report(result: Dict[str, Any]) -> BytesIO:
    """
    导出预检报告
    返回 BytesIO 对象，可以直接通过 FastAPI 返回
    """
    wb = Workbook()
    # 删除默认的工作表
    wb.remove(wb.active)

    # 1. 汇总表
    summary = result.get("summary", {})
    add_summary_sheet(wb, summary)

    # 2. 错误库命中
    headers = ["行号", "姓名", "身份证号", "所在村", "所在组", "错误类型", "错误原因", "来源"]
    data = [
        {"行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
         "所在村": r.get("village"), "所在组": r.get("group"),
         "错误类型": r.get("error_type"), "错误原因": r.get("error_reason"), "来源": r.get("source")}
        for r in result.get("error_library_hits", [])
    ]
    add_sheet_from_data(wb, "错误库命中", headers, data)

    # 3. 格式错误
    headers = ["行号", "姓名", "身份证号", "所在村", "所在组", "错误内容"]
    data = [
        {"行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
         "所在村": r.get("village"), "所在组": r.get("group"),
         "错误内容": "；".join(r["errors"]) if isinstance(r.get("errors"), list) else r.get("errors", "")}
        for r in result.get("format_errors", [])
    ]
    add_sheet_from_data(wb, "格式错误", headers, data)

    # 4. 村组不存在
    headers = ["行号", "姓名", "身份证号", "所在村", "所在组", "错误信息"]
    data = [
        {"行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
         "所在村": r.get("village"), "所在组": r.get("group"), "错误信息": r.get("error")}
        for r in result.get("village_errors", [])
    ]
    add_sheet_from_data(wb, "村组不存在", headers, data)

    # 5. 重复身份证
    headers = ["行号", "姓名", "身份证号", "错误信息"]
    data = [
        {"行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"), "错误信息": r.get("error")}
        for r in result.get("duplicate_errors", [])
    ]
    add_sheet_from_data(wb, "重复身份证", headers, data)

    # 5.5 数据库已有申请记录
    headers = ["行号", "姓名", "身份证号", "所在村", "所在组", "已有申请记录", "错误信息"]
    data = [
        {"行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
         "所在村": r.get("village"), "所在组": r.get("group"),
         "已有申请记录": r.get("existing_apps"), "错误信息": r.get("error")}
        for r in result.get("db_duplicate_apps", [])
    ]
    add_sheet_from_data(wb, "数据库已有申请记录", headers, data)

    # 6. 性别不符
    headers = ["行号", "姓名", "身份证号", "Excel性别", "身份证性别"]
    data = [
        {"行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
         "Excel性别": r.get("excel_gender"), "身份证性别": r.get("id_card_gender")}
        for r in result.get("gender_mismatch", [])
    ]
    add_sheet_from_data(wb, "性别不符", headers, data)

    # 7. 面积异常
    headers = [
        "行号", "姓名", "身份证号", "所在村", "所在组",
        "异常类型", "异常详情",
        "Excel承包地面积", "数据库承包面积", "流转出面积", "代耕代种进",
        "不补贴面积", "实际补贴面积", "自有承包地占用", "户级当季已有申请",
        "户级合计", "超出面积"
    ]
    data = []
    for r in result.get("area_anomalies", []):
        data.append({
            "行号": r.get("row"),
            "姓名": r.get("name"),
            "身份证号": r.get("id_card"),
            "所在村": r.get("village"),
            "所在组": r.get("group"),
            "异常类型": r.get("anomaly_type"),
            "异常详情": r.get("anomaly_details"),
            "Excel承包地面积": r.get("contract_area"),
            "数据库承包面积": r.get("db_contract_area"),
            "流转出面积": r.get("trust_out_area"),
            "代耕代种进": r.get("trust_in_area"),
            "不补贴面积": r.get("no_subsidy_area"),
            "实际补贴面积": r.get("actual_subsidy_area"),
            "自有承包地占用": r.get("self_occupy"),
            "户级当季已有申请": r.get("hh_used"),
            "户级合计": r.get("hh_total"),
            "超出面积": r.get("exceed_amount"),
        })
    add_sheet_from_data(wb, "面积异常", headers, data)

    # 8. 同一家庭多成员申请
    headers = [
        "行号", "姓名", "身份证号", "所在村", "所在组",
        "家庭户ID", "申请人数", "其他成员", "Excel备注", "数据库已有申请记录备注"
    ]
    data = []
    for r in result.get("household_duplicates", []):
        # 格式化数据库已有申请记录备注
        db_remarks = []
        for app in r.get("db_existing_apps", []):
            remark_text = f"{app.get('real_name', '')}({app.get('subsidy_name', '')})"
            if app.get('remark'):
                remark_text += f": {app['remark']}"
            db_remarks.append(remark_text)
        db_remark_str = "；".join(db_remarks) if db_remarks else ""

        data.append({
            "行号": r.get("row"),
            "姓名": r.get("name"),
            "身份证号": r.get("id_card"),
            "所在村": r.get("village"),
            "所在组": r.get("group"),
            "家庭户ID": r.get("household_id"),
            "申请人数": r.get("total_count"),
            "其他成员": "、".join(r.get("other_members", [])) if r.get("other_members") else "",
            "Excel备注": r.get("excel_remark", ""),
            "数据库已有申请记录备注": db_remark_str,
        })
    add_sheet_from_data(wb, "同一家庭多成员申请", headers, data)

    # 10. 新增农户
    headers = ["行号", "姓名", "身份证号", "所在村", "所在组", "说明"]
    data = []
    for r in result.get("new_farmers", []):
        data.append({
            "行号": r.get("row"),
            "姓名": r.get("name"),
            "身份证号": r.get("id_card"),
            "所在村": r.get("village"),
            "所在组": r.get("group"),
            "说明": "数据库中不存在，将新增",
        })
    add_sheet_from_data(wb, "新增农户", headers, data)

    # 11. 减少农户
    headers = ["姓名", "身份证号", "所在村", "所在组", "说明"]
    data = [
        {"姓名": r.get("name"), "身份证号": r.get("id_card"),
         "所在村": r.get("village"), "所在组": r.get("group"),
         "说明": "有补贴记录但未在本次导入中"}
        for r in result.get("removed_farmers", [])
    ]
    add_sheet_from_data(wb, "减少农户", headers, data)

    # 12. 字段变更
    headers = ["行号", "姓名", "身份证号", "变更内容"]
    data = [
        {"行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
         "变更内容": "；".join(r["changes"]) if isinstance(r.get("changes"), list) else r.get("changes", "")}
        for r in result.get("changed_farmers", [])
    ]
    add_sheet_from_data(wb, "字段变更", headers, data)

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_confirmed_area_diff(rows: List[Dict[str, Any]]) -> BytesIO:
    """
    导出确权面积与承包面积对比报告。

    rows 每条包含：
      household_name, village_full_name, head_name,
      contract_area, confirmed_area, diff, status, label
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("确权面积对比")
    headers = [
        "家庭户", "所在村组", "户主姓名",
        "承包面积(亩)", "确权面积(亩)", "差值(确权-承包)",
        "对比结果",
    ]
    ws.append(headers)

    # 状态对应填充色
    STATUS_FILL = {
        "match":            PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        "confirmed_larger": PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid"),
        "contract_larger":  PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid"),
        "missing":          PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
    }

    for row in rows:
        status = row.get("status", "missing")
        ws.append([
            row.get("household_name", ""),
            row.get("village_full_name", ""),
            row.get("head_name", ""),
            row.get("contract_area", ""),
            row.get("confirmed_area", ""),
            row.get("diff", ""),
            row.get("label", ""),
        ])
        # 最后一行上色
        fill = STATUS_FILL.get(status, PatternFill())
        for cell in ws[ws.max_row]:
            cell.fill = fill

    apply_styles(ws, has_alternate_rows=False)
    # 重新给表头加回蓝色（apply_styles 已覆盖）
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _filter_data_by_village(data: List[Dict[str, Any]], village_name: str) -> List[Dict[str, Any]]:
    """按村筛选数据"""
    result = []
    for row in data:
        row_village = row.get("village", row.get("village_name", ""))
        if row_village == village_name:
            result.append(row)
    return result


def _transform_sheet_data(sheet_name: str, raw_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """将原始英文key数据转换为中文key数据，供 add_sheet_from_data 使用"""

    def join_list(val, sep="；"):
        if isinstance(val, (list, tuple)):
            return sep.join(str(x) for x in val)
        return val or ""

    transforms = {
        "错误库命中": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "所在村": r.get("village"), "所在组": r.get("group"),
            "错误类型": r.get("error_type"), "错误原因": r.get("error_reason"), "来源": r.get("source"),
        },
        "格式错误": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "所在村": r.get("village"), "所在组": r.get("group"),
            "错误内容": join_list(r.get("errors")),
        },
        "村组不存在": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "所在村": r.get("village"), "所在组": r.get("group"), "错误信息": r.get("error"),
        },
        "重复身份证": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "错误信息": r.get("error"),
        },
        "性别不符": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "Excel性别": r.get("excel_gender"), "身份证性别": r.get("id_card_gender"),
        },
        "面积异常": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "所在村": r.get("village"), "所在组": r.get("group"),
            "异常类型": r.get("anomaly_type"), "异常详情": r.get("anomaly_details"),
            "Excel承包地面积": r.get("contract_area"), "数据库承包面积": r.get("db_contract_area"),
            "流转出面积": r.get("trust_out_area"), "代耕代种进": r.get("trust_in_area"),
            "不补贴面积": r.get("no_subsidy_area"), "实际补贴面积": r.get("actual_subsidy_area"),
            "自有承包地占用": r.get("self_occupy"), "户级当季已有申请": r.get("hh_used"),
            "户级合计": r.get("hh_total"), "超出面积": r.get("exceed_amount"),
        },
        "承包面积缺失": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "所在村": r.get("village"), "所在组": r.get("group"),
            "申请面积": r.get("apply_area", r.get("contract_area")),
            "说明": r.get("reason", r.get("note", "")),
        },
        "年龄异常": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "所在村": r.get("village"), "所在组": r.get("group"),
            "年龄": r.get("age"), "出生年份": r.get("birth_year"),
            "说明": r.get("reason", r.get("note", "")),
        },
        "死亡农户": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "所在村": r.get("village"), "所在组": r.get("group"),
            "说明": r.get("reason", r.get("note", "")),
        },
        "同一家庭多成员申请": lambda r: {
            "行号": r.get("row"),
            "姓名": r.get("name"),
            "身份证号": r.get("id_card"),
            "所在村": r.get("village"),
            "所在组": r.get("group"),
            "家庭户ID": r.get("household_id"),
            "申请人数": r.get("total_count"),
            "其他成员": "、".join(r.get("other_members", [])) if r.get("other_members") else "",
            "Excel备注": r.get("excel_remark", ""),
            "数据库已有申请记录备注": "；".join([
                f"{app.get('real_name', '')}({app.get('subsidy_name', '')})" +
                (f": {app['remark']}" if app.get('remark') else "")
                for app in r.get("db_existing_apps", [])
            ]) if r.get("db_existing_apps") else "",
        },
        "新增农户": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "所在村": r.get("village"), "所在组": r.get("group"),
            "说明": "数据库中不存在，将新增",
        },
        "减少农户": lambda r: {
            "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "所在村": r.get("village"), "所在组": r.get("group"),
            "说明": "有补贴记录但未在本次导入中",
        },
        "字段变更": lambda r: {
            "行号": r.get("row"), "姓名": r.get("name"), "身份证号": r.get("id_card"),
            "变更内容": join_list(r.get("changes")),
        },
    }

    transform_fn = transforms.get(sheet_name)
    if not transform_fn:
        return raw_rows
    return [transform_fn(row) for row in raw_rows]


def _get_all_villages(result: Dict[str, Any]) -> List[str]:
    """从结果中获取所有涉及的村名"""
    villages = set()

    # 遍历所有可能包含村信息的字段
    village_fields = ["error_library_hits", "format_errors", "village_errors",
                      "duplicate_errors", "db_duplicate_apps", "gender_mismatch", "area_anomalies",
                      "new_farmers", "removed_farmers", "changed_farmers",
                      "household_duplicates"
                      ]

    for field in village_fields:
        data = result.get(field, [])
        for row in data:
            village = row.get("village", row.get("village_name", ""))
            if village:
                villages.add(village)

    return sorted(list(villages))


def export_precheck_report_by_village(
    result: Dict[str, Any],
    village_name: str,
    selected_sheets: Optional[List[str]] = None
) -> BytesIO:
    """
    导出单个村的预检查报告
    """
    wb = Workbook()
    wb.remove(wb.active)

    sheet_configs = {
        "汇总": (add_summary_sheet, "summary"),
        "错误库命中": (add_sheet_from_data, "error_library_hits",
                       ["行号", "姓名", "身份证号", "所在村", "所在组", "错误类型", "错误原因", "来源"]),
        "格式错误": (add_sheet_from_data, "format_errors",
                    ["行号", "姓名", "身份证号", "所在村", "所在组", "错误内容"]),
        "村组不存在": (add_sheet_from_data, "village_errors",
                     ["行号", "姓名", "身份证号", "所在村", "所在组", "错误信息"]),
        "重复身份证": (add_sheet_from_data, "duplicate_errors",
                      ["行号", "姓名", "身份证号", "错误信息"]),
        "数据库已有申请记录": (add_sheet_from_data, "db_duplicate_apps",
                              ["行号", "姓名", "身份证号", "所在村", "所在组", "已有申请记录", "错误信息"]),
        "性别不符": (add_sheet_from_data, "gender_mismatch",
                   ["行号", "姓名", "身份证号", "Excel性别", "身份证性别"]),
        "面积异常": (add_sheet_from_data, "area_anomalies",
                   ["行号", "姓名", "身份证号", "所在村", "所在组", "异常类型", "异常详情",
                    "Excel承包地面积", "数据库承包面积", "流转出面积", "代耕代种进",
                    "不补贴面积", "实际补贴面积", "自有承包地占用", "户级当季已有申请",
                    "户级合计", "超出面积"]),
        "承包面积缺失": (add_sheet_from_data, "area_missing",
                     ["行号", "姓名", "身份证号", "所在村", "所在组", "申请面积", "说明"]),
        "年龄异常": (add_sheet_from_data, "age_anomaly",
                   ["行号", "姓名", "身份证号", "所在村", "所在组", "年龄", "出生年份", "说明"]),
        "死亡农户": (add_sheet_from_data, "deceased_farmers",
                   ["行号", "姓名", "身份证号", "所在村", "所在组", "说明"]),
        "受限身份农户": (add_sheet_from_data, "restricted_farmers",
                      ["行号", "姓名", "身份证号", "所在村", "所在组", "说明"]),
        "同一家庭多成员申请": (add_sheet_from_data, "household_duplicates",
                           ["行号", "姓名", "身份证号", "所在村", "所在组",
                            "家庭户ID", "申请人数", "其他成员", "Excel备注", "数据库已有申请记录备注"]),
        "新增农户": (add_sheet_from_data, "new_farmers",
                    ["行号", "姓名", "身份证号", "所在村", "所在组", "说明"]),
        "减少农户": (add_sheet_from_data, "removed_farmers",
                    ["姓名", "身份证号", "所在村", "所在组", "说明"]),
        "字段变更": (add_sheet_from_data, "changed_farmers",
                    ["行号", "姓名", "身份证号", "变更内容"]),
    }

    # 如果没有指定sheet，默认全部
    if not selected_sheets:
        selected_sheets = list(sheet_configs.keys())

    for sheet_name in selected_sheets:
        if sheet_name not in sheet_configs:
            continue

        config = sheet_configs[sheet_name]
        if sheet_name == "汇总":
            # 汇总表需要特殊处理，只显示该村相关的统计
            add_summary_sheet(wb, result.get("summary", {}))
        else:
            func, data_key, headers = config
            data = result.get(data_key, [])
            filtered_data = _filter_data_by_village(data, village_name)
            transformed = _transform_sheet_data(sheet_name, filtered_data)
            if transformed:
                add_sheet_from_data(wb, sheet_name, headers, transformed)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_precheck_report_zip(
    result: Dict[str, Any],
    selected_sheets: Optional[List[str]] = None,
    file_name: str = "预检查报告"
) -> BytesIO:
    """
    导出分村的预检查报告，打包成 ZIP
    每个村一个 Excel 文件
    """
    from datetime import datetime

    zip_buffer = BytesIO()

    # 获取所有涉及的村
    villages = _get_all_villages(result)

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # 为每个村生成 Excel 并添加到 ZIP
        for village in villages:
            excel_buffer = export_precheck_report_by_village(result, village, selected_sheets)
            date_str = datetime.now().strftime("%Y-%m-%d")
            safe_village_name = village.replace("/", "_").replace("\\", "_")
            excel_filename = f"{file_name}_{safe_village_name}_{date_str}.xlsx"
            zipf.writestr(excel_filename, excel_buffer.getvalue())

        # 如果没有村的数据，生成一个说明文件
        if not villages:
            zipf.writestr("说明.txt", "本次检查数据中没有村信息，无法按村拆分导出。")

    zip_buffer.seek(0)
    return zip_buffer


def export_precheck_report_with_options(
    result: Dict[str, Any],
    split_by_village: bool = False,
    selected_sheets: Optional[List[str]] = None,
    file_name: str = "预检查报告"
) -> tuple[BytesIO, str, str]:
    """
    根据选项导出预检查报告

    Returns:
        (buffer, filename, media_type)
    """
    from datetime import datetime

    date_str = datetime.now().strftime("%Y-%m-%d")

    if split_by_village:
        # 分村导出，返回 ZIP
        buffer = export_precheck_report_zip(result, selected_sheets, file_name)
        filename = f"{file_name}_分村_{date_str}.zip"
        media_type = "application/zip"
    else:
        # 不分村，按选择的 sheet 导出单个 Excel
        wb = Workbook()
        wb.remove(wb.active)

        # sheet 名称 → (result key, headers)，None 表示汇总特殊处理
        sheet_defs = [
            ("汇总", None, None),
            ("错误库命中", "error_library_hits",
             ["行号", "姓名", "身份证号", "所在村", "所在组", "错误类型", "错误原因", "来源"]),
            ("格式错误", "format_errors",
             ["行号", "姓名", "身份证号", "所在村", "所在组", "错误内容"]),
            ("村组不存在", "village_errors",
             ["行号", "姓名", "身份证号", "所在村", "所在组", "错误信息"]),
            ("重复身份证", "duplicate_errors",
             ["行号", "姓名", "身份证号", "错误信息"]),
            ("性别不符", "gender_mismatch",
             ["行号", "姓名", "身份证号", "Excel性别", "身份证性别"]),
            ("面积异常", "area_anomalies",
             ["行号", "姓名", "身份证号", "所在村", "所在组", "异常类型", "异常详情",
              "Excel承包地面积", "数据库承包面积", "流转出面积", "代耕代种进",
              "不补贴面积", "实际补贴面积", "自有承包地占用", "户级当季已有申请",
              "户级合计", "超出面积"]),
            ("承包面积缺失", "area_missing",
             ["行号", "姓名", "身份证号", "所在村", "所在组", "申请面积", "说明"]),
            ("年龄异常", "age_anomaly",
             ["行号", "姓名", "身份证号", "所在村", "所在组", "年龄", "出生年份", "说明"]),
            ("死亡农户", "deceased_farmers",
             ["行号", "姓名", "身份证号", "所在村", "所在组", "说明"]),
            ("受限身份农户", "restricted_farmers",
             ["行号", "姓名", "身份证号", "所在村", "所在组", "说明"]),
            ("同一家庭多成员申请", "household_duplicates",
             ["行号", "姓名", "身份证号", "所在村", "所在组",
              "家庭户ID", "申请人数", "其他成员", "Excel备注", "数据库已有申请记录备注"]),
            ("新增农户", "new_farmers",
             ["行号", "姓名", "身份证号", "所在村", "所在组", "说明"]),
            ("减少农户", "removed_farmers",
             ["姓名", "身份证号", "所在村", "所在组", "说明"]),
            ("字段变更", "changed_farmers",
             ["行号", "姓名", "身份证号", "变更内容"]),
        ]

        # 如果没有选择 sheet，默认全部
        if not selected_sheets:
            selected_sheets = [name for name, _, _ in sheet_defs]

        # 按顺序添加 sheet
        for sheet_name, data_key, headers in sheet_defs:
            if sheet_name not in selected_sheets:
                continue
            if sheet_name == "汇总":
                add_summary_sheet(wb, result.get("summary", {}))
            else:
                raw_data = result.get(data_key, [])
                transformed = _transform_sheet_data(sheet_name, raw_data)
                add_sheet_from_data(wb, sheet_name, headers, transformed)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"{file_name}_{date_str}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return buffer, filename, media_type

